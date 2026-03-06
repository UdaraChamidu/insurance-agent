import csv
import io
from typing import List

import mammoth
from openpyxl import load_workbook
from pdfminer.high_level import extract_text

class DocumentProcessor:
    def __init__(self):
        pass

    def process_txt(self, content: bytes) -> str:
        return content.decode("utf-8", errors="ignore")

    def process_md(self, content: bytes) -> str:
        # Markdown stays as plain text for structural chunking downstream.
        return content.decode("utf-8", errors="ignore")

    def process_csv(self, content: bytes) -> str:
        try:
            decoded = content.decode("utf-8", errors="ignore")
            reader = csv.reader(io.StringIO(decoded))
            lines: List[str] = []
            for row in reader:
                cleaned = [str(cell or "").strip() for cell in row]
                if not any(cleaned):
                    continue
                lines.append(" | ".join(cleaned))
            return "\n".join(lines)
        except Exception as e:
            print(f"Error processing CSV: {e}")
            return ""

    async def process_xlsx(self, content: bytes) -> str:
        try:
            import asyncio

            def _extract(c: bytes) -> str:
                wb = load_workbook(io.BytesIO(c), read_only=True, data_only=True)
                sheet_blocks: List[str] = []
                for sheet in wb.worksheets:
                    rows: List[str] = [f"# Sheet: {sheet.title}"]
                    for row in sheet.iter_rows(values_only=True):
                        cells = [str(cell).strip() if cell is not None else "" for cell in row]
                        if not any(cells):
                            continue
                        rows.append(" | ".join(cells))
                    if len(rows) > 1:
                        sheet_blocks.append("\n".join(rows))
                try:
                    wb.close()
                except Exception:
                    pass
                return "\n\n".join(sheet_blocks)

            loop = asyncio.get_event_loop()
            return await loop.run_in_executor(None, _extract, content)
        except Exception as e:
            print(f"Error processing XLSX: {e}")
            return ""

    async def process_pdf(self, content: bytes) -> str:
        try:
            # pdfminer.six synchronous call, wrap in thread if cpu bound?
            # It's CPU bound, but for small files ok.
            # Best practice: run_in_executor
            import asyncio
            loop = asyncio.get_event_loop()
            text = await loop.run_in_executor(None, extract_text, io.BytesIO(content))
            return text
        except Exception as e:
            print(f"Error processing PDF: {e}")
            return ""
    
    async def process_docx(self, content: bytes) -> str:
        try:
            # mammoth is sync
            import asyncio
            loop = asyncio.get_event_loop()
            
            def _extract(c):
                result = mammoth.extract_raw_text(io.BytesIO(c))
                return result.value
                
            text = await loop.run_in_executor(None, _extract, content)
            return text
        except Exception as e:
            print(f"Error processing DOCX: {e}")
            return ""

document_processor = DocumentProcessor()
